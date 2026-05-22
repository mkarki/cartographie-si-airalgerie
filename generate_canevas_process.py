"""
Genere le canevas Excel a joindre au mail de recueil des process cle utilisateurs.

Produit : CANEVAS_PROCESS_KEY_USERS.xlsx
Structure : en-tete projet + 3 colonnes (Process / Systemes / Outputs-KPI) + lignes vides.
"""
import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

OUT = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'CANEVAS_PROCESS_KEY_USERS.xlsx')


def main():
    wb = Workbook()

    # ---- Feuille 1 : Canevas ----
    ws = wb.active
    ws.title = 'Process'

    # Couleurs Air Algerie (rouge officiel + beige doux)
    AH_RED = 'C8102E'
    AH_RED_LIGHT = 'F4B3BC'
    HEADER_FILL = PatternFill(start_color=AH_RED, end_color=AH_RED, fill_type='solid')
    SUB_FILL = PatternFill(start_color='F5F5F5', end_color='F5F5F5', fill_type='solid')
    EXAMPLE_FILL = PatternFill(start_color='FFF4E6', end_color='FFF4E6', fill_type='solid')

    thin = Side(border_style='thin', color='B0B0B0')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    title_font = Font(name='Calibri', size=14, bold=True, color='FFFFFF')
    header_font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
    note_font = Font(name='Calibri', size=10, italic=True, color='555555')
    data_font = Font(name='Calibri', size=11)

    # Ligne 1 : titre projet
    ws.merge_cells('A1:E1')
    ws['A1'] = 'Air Algerie — Cartographie SI — Recueil des process metier'
    ws['A1'].font = title_font
    ws['A1'].fill = HEADER_FILL
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 28

    # Ligne 2 : identification key user
    ws['A2'] = 'Key User :'
    ws['A2'].font = Font(bold=True)
    ws.merge_cells('B2:E2')
    ws['B2'] = ''  # a remplir par le destinataire

    ws['A3'] = 'Division / Direction :'
    ws['A3'].font = Font(bold=True)
    ws.merge_cells('B3:E3')
    ws['B3'] = ''

    ws['A4'] = 'Date :'
    ws['A4'].font = Font(bold=True)
    ws.merge_cells('B4:E4')
    ws['B4'] = ''

    for row in (2, 3, 4):
        ws[f'A{row}'].fill = SUB_FILL
        ws[f'B{row}'].fill = SUB_FILL
        ws[f'A{row}'].border = border
        ws[f'B{row}'].border = border

    # Ligne 6 : note
    ws.merge_cells('A6:E6')
    ws['A6'] = ("Merci de lister ci-dessous les process metier relevant de votre perimetre. "
                "Pas de description detaillee necessaire : nous reviendrons vers vous si un approfondissement est requis.")
    ws['A6'].font = note_font
    ws['A6'].alignment = Alignment(wrap_text=True, vertical='center')
    ws.row_dimensions[6].height = 30

    # Ligne 8 : en-tetes du tableau
    headers = ['N°', 'Nom du process', 'Description', 'Systemes utilises', 'Outputs / KPI produits']
    for col_idx, h in enumerate(headers, start=1):
        cell = ws.cell(row=8, column=col_idx, value=h)
        cell.font = header_font
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = border
    ws.row_dimensions[8].height = 30

    # Ligne 9 : exemple grise
    example = [
        '(ex.)',
        'Programmation des vols hebdomadaire',
        "Construction et publication du planning de vols de la semaine N+1, integrant rotations equipages et appareils.",
        'AIMS, Excel',
        'Fichier SSIM hebdo, taux de respect planning',
    ]
    for col_idx, val in enumerate(example, start=1):
        cell = ws.cell(row=9, column=col_idx, value=val)
        cell.font = Font(italic=True, color='666666')
        cell.fill = EXAMPLE_FILL
        cell.alignment = Alignment(wrap_text=True, vertical='center')
        cell.border = border
    ws.row_dimensions[9].height = 30

    # Lignes 10-29 : 20 lignes vides numerotees
    for i, row in enumerate(range(10, 30), start=1):
        ws.cell(row=row, column=1, value=i).alignment = Alignment(horizontal='center', vertical='center')
        ws.cell(row=row, column=1).border = border
        ws.cell(row=row, column=1).font = data_font
        for col_idx in (2, 3, 4, 5):
            cell = ws.cell(row=row, column=col_idx, value='')
            cell.border = border
            cell.alignment = Alignment(wrap_text=True, vertical='center')
            cell.font = data_font
        ws.row_dimensions[row].height = 28

    # Largeur des colonnes
    ws.column_dimensions['A'].width = 6
    ws.column_dimensions['B'].width = 38
    ws.column_dimensions['C'].width = 55
    ws.column_dimensions['D'].width = 32
    ws.column_dimensions['E'].width = 40

    # Figer la ligne d'en-tete
    ws.freeze_panes = 'A9'

    # ---- Feuille 2 : Instructions ----
    ws2 = wb.create_sheet('Instructions')
    ws2.column_dimensions['A'].width = 110

    lines = [
        ('Canevas — Recueil des process metier', title_font, HEADER_FILL, 28),
        ('', None, None, 10),
        ('Objectif', Font(bold=True, size=12), None, 22),
        ("Recueillir la liste des process metier de chaque division pour les integrer dans la cartographie du SI "
         "et les relier aux systemes informatiques utilises.", data_font, None, 40),
        ('', None, None, 10),
        ('Comment remplir', Font(bold=True, size=12), None, 22),
        ("1. Renseignez vos informations en haut de la feuille 'Process' (nom, division, date).", data_font, None, 22),
        ("2. Listez chaque process sur une ligne, avec :", data_font, None, 22),
        ("   - le nom du process (ex. 'Traitement des demandes de remboursement passagers')", data_font, None, 22),
        ("   - une description courte en 1-2 phrases (ex. 'Reception des demandes via formulaire web, controle pieces, validation et remboursement')", data_font, None, 22),
        ("   - le ou les systemes utilises (ex. 'RAPID PASSAGERS, Sage Finance, Excel')", data_font, None, 22),
        ("   - les outputs ou KPI produits (ex. 'Rapport mensuel des remboursements, delai moyen de traitement')", data_font, None, 22),
        ("3. Pas besoin d'etre exhaustif sur la description : nous reviendrons vers vous si besoin.", data_font, None, 22),
        ("4. Renvoyez le fichier rempli par mail avant la date demandee.", data_font, None, 22),
        ('', None, None, 10),
        ('Questions / contact', Font(bold=True, size=12), None, 22),
        ("Pour toute question sur le remplissage : Mohamedamine KARKI — equipe projet Cartographie SI.", data_font, None, 22),
    ]
    for i, (text, font, fill, height) in enumerate(lines, start=1):
        cell = ws2.cell(row=i, column=1, value=text)
        if font:
            cell.font = font
        if fill:
            cell.fill = fill
        cell.alignment = Alignment(wrap_text=True, vertical='center')
        ws2.row_dimensions[i].height = height

    wb.save(OUT)
    size_kb = os.path.getsize(OUT) / 1024
    print(f'OK  Canevas genere : {OUT} ({size_kb:.1f} KB)')


if __name__ == '__main__':
    main()
