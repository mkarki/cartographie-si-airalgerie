"""
Genere LISTE_KEY_USERS.xlsx : tableau Nom/Prenom (rempli) + Email (a remplir
par la secretaire) + Systeme(s) + Direction + Role.

Source : base cartography_questionnaire (noms complets) + EMAILS_KEY_USERS.md
         (pour la liste des 62 key users uniques et leur rattachement direction).
Sortie : LISTE_KEY_USERS.xlsx
"""
import os
import re
import sqlite3
from collections import OrderedDict

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

BASE = os.path.dirname(os.path.abspath(__file__))
SRC_MD = os.path.join(BASE, 'EMAILS_KEY_USERS.md')
DB = os.path.join(BASE, 'db.sqlite3')
OUT = os.path.join(BASE, 'LISTE_KEY_USERS.xlsx')


def parse_md():
    """Renvoie liste ordonnee : [ { 'name', 'email', 'role', 'direction', 'systems':set } ]
    deduplique par email.
    """
    with open(SRC_MD, 'r', encoding='utf-8') as f:
        text = f.read().split('## Récapitulatif')[0]

    by_email = OrderedDict()
    current_dir = None
    current_sys = None

    for line in text.split('\n'):
        m_dir = re.match(r'^##\s+([A-Z]{2,5})\s+—', line)
        if m_dir:
            current_dir = m_dir.group(1)
            continue
        m_sys = re.match(r'^###\s+(.+?)\s*$', line)
        if m_sys:
            current_sys = re.sub(r'\s*\*\(.*?\)\*\s*$', '', m_sys.group(1)).strip()
            continue
        m_row = re.match(r'^\|\s*(Key User[^|]*?)\s*\|\s*([^|]+?)\s*\|\s*([^|]+?)\s*\|\s*$', line)
        if m_row and current_dir and current_sys:
            role = m_row.group(1).strip()
            name = m_row.group(2).strip()
            email = m_row.group(3).strip()
            if email in ('—', '-', '') or 'Non désigné' in name:
                continue
            if email not in by_email:
                by_email[email] = {
                    'name': name,
                    'email_actuel': email,
                    'directions': OrderedDict(),
                    'systems': [],
                    'roles': set(),
                }
            by_email[email]['directions'][current_dir] = True
            if current_sys not in by_email[email]['systems']:
                by_email[email]['systems'].append(current_sys)
            if 'Principal' in role:
                by_email[email]['roles'].add('Principal')
            elif 'Backup' in role:
                by_email[email]['roles'].add('Backup')
    return by_email


def get_full_names_from_db():
    """Retourne un dict { nom_court_upper : nom_complet } extrait de la cartographie."""
    conn = sqlite3.connect(DB)
    cur = conn.cursor()
    cur.execute('SELECT key_users, key_users_backup FROM cartography_questionnaire')
    fullnames = {}
    for principals, backups in cur.fetchall():
        for raw in (principals or '') + ',' + (backups or ''):
            pass
        for chunk in re.split(r'[,/]', (principals or '') + ',' + (backups or '')):
            # Nettoyer '(DOS)', '(DRM)' etc.
            c = re.sub(r'\s*\([^)]*\)\s*', '', chunk).strip()
            if not c or c == '—':
                continue
            # Cle = 1 ou 2 premiers tokens en MAJ (nom de famille)
            tokens = c.split()
            if not tokens:
                continue
            key1 = tokens[0].upper()
            key2 = ' '.join(tokens[:2]).upper() if len(tokens) > 1 else key1
            for k in {key1, key2}:
                if k not in fullnames or len(fullnames[k]) < len(c):
                    fullnames[k] = c
    conn.close()
    return fullnames


def best_fullname(name_md: str, fullnames: dict) -> str:
    """Essaie de trouver le nom complet (avec prenom) dans la base."""
    # Retirer les civilites M./Mme/Mlle
    n = re.sub(r'^(M\.?|Mme\.?|Mlle\.?)\s+', '', name_md).strip()
    tokens = n.split()
    if not tokens:
        return name_md
    # Cles possibles, de la plus specifique a la plus generale
    candidates = []
    if len(tokens) >= 2:
        candidates.append(' '.join(tokens[:2]).upper())
    candidates.append(tokens[0].upper())
    for c in candidates:
        if c in fullnames:
            full = fullnames[c]
            # Si on a au moins autant de tokens, prendre la version DB
            if len(full.split()) > len(n.split()):
                return full
            if len(full) >= len(n):
                return full
    return n


def build_xlsx(users, fullnames):
    wb = Workbook()
    ws = wb.active
    ws.title = 'Key Users'

    # Styles
    header_fill = PatternFill('solid', fgColor='1F4E78')
    header_font = Font(bold=True, color='FFFFFF', size=11)
    todo_fill = PatternFill('solid', fgColor='FFF2CC')  # jaune clair
    thin = Side(border_style='thin', color='999999')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    center = Alignment(horizontal='center', vertical='center')
    left = Alignment(horizontal='left', vertical='center', wrap_text=True)

    headers = ['#', 'Nom & Prénom', 'Email (à compléter)', 'Direction', 'Rôle', 'Système(s) couvert(s)']
    ws.append(headers)
    for c, _ in enumerate(headers, 1):
        cell = ws.cell(row=1, column=c)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center
        cell.border = border

    for i, (_, info) in enumerate(users.items(), 1):
        full = best_fullname(info['name'], fullnames)
        directions = ' / '.join(info['directions'].keys())
        role = 'Principal' if 'Principal' in info['roles'] else 'Backup'
        if info['roles'] == {'Principal', 'Backup'}:
            role = 'Principal + Backup'
        systems = ', '.join(info['systems'])
        ws.append([i, full, '', directions, role, systems])

        for c in range(1, 7):
            cell = ws.cell(row=i + 1, column=c)
            cell.border = border
            cell.alignment = left if c in (2, 6) else center
        # Marquer la colonne Email en jaune
        ws.cell(row=i + 1, column=3).fill = todo_fill

    # Largeurs
    widths = [5, 32, 40, 14, 18, 70]
    for idx, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(idx)].width = w

    ws.row_dimensions[1].height = 28
    ws.freeze_panes = 'A2'

    # Note dans une seconde feuille
    ws2 = wb.create_sheet('Note')
    ws2['A1'] = "Consigne"
    ws2['A1'].font = Font(bold=True, size=12)
    ws2['A2'] = ("Merci de compléter la colonne « Email (à compléter) » pour chaque key user.\n"
                 "Les lignes sans prénom (nom suivi d'une initiale ou d'un seul mot) sont à compléter "
                 "si possible.")
    ws2['A2'].alignment = Alignment(wrap_text=True, vertical='top')
    ws2.column_dimensions['A'].width = 100
    ws2.row_dimensions[2].height = 60

    wb.save(OUT)
    print(f'OK  {OUT}')
    print(f'    {len(users)} lignes')


def main():
    users = parse_md()
    fullnames = get_full_names_from_db()
    build_xlsx(users, fullnames)


if __name__ == '__main__':
    main()
