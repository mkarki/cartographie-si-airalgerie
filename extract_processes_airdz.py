"""
Extrait les process des canevas Excel remplis par les key users
(dossier 'process AIRDZ/') et produit un JSON consolidé.

Usage:
    python extract_processes_airdz.py

Sortie:
    process_airdz_extracted.json (à la racine du repo)
"""
import json
import os
import glob

import openpyxl


SOURCE_DIR = '/Users/mohamedamine/Air Algérie/process AIRDZ'
OUT_FILE = os.path.join(os.path.dirname(os.path.abspath(__file__)),
                        'process_airdz_extracted.json')


def cv(x):
    if x is None:
        return ''
    if isinstance(x, str):
        return x.strip()
    return str(x).strip()


def parse_workbook(path):
    rel = os.path.relpath(path, SOURCE_DIR)
    direction_folder = rel.split(os.sep)[0]
    wb = openpyxl.load_workbook(path, data_only=True)
    sheet_name = 'Process' if 'Process' in wb.sheetnames else wb.sheetnames[0]
    ws = wb[sheet_name]
    rows = [list(r) for r in ws.iter_rows(values_only=True)]

    key_user = ''
    division = ''
    date = ''
    for r in rows[:8]:
        if not r:
            continue
        first = cv(r[0])
        second = next((cv(c) for c in r[1:] if cv(c)), '')
        if 'Key User' in first:
            key_user = second
        elif 'Division' in first or 'Direction' in first:
            division = second
        elif first.startswith('Date'):
            date = second[:10]

    header_idx = None
    for i, r in enumerate(rows):
        if r and cv(r[0]) == 'N°':
            header_idx = i
            break

    processes = []
    if header_idx is not None:
        for r in rows[header_idx + 1:]:
            if not r:
                continue
            num_raw = cv(r[0])
            name = cv(r[1] if len(r) > 1 else '')
            description = cv(r[2] if len(r) > 2 else '')
            systems = cv(r[3] if len(r) > 3 else '')
            kpi = cv(r[4] if len(r) > 4 else '')
            if not name or 'ex' in num_raw.lower()[:3]:
                continue
            processes.append({
                'num': num_raw,
                'name': name,
                'description': description,
                'systems_raw': systems,
                'outputs_kpi': kpi,
            })

    return {
        'file': rel,
        'direction_folder': direction_folder,
        'key_user': key_user,
        'division': division,
        'date': date,
        'processes': processes,
    }


def main():
    patterns = ['**/*.xlsx', '**/*.XLSX']
    files = []
    for p in patterns:
        files.extend(glob.glob(os.path.join(SOURCE_DIR, p), recursive=True))
    files = sorted(set(files))

    out = []
    for f in files:
        out.append(parse_workbook(f))

    with open(OUT_FILE, 'w', encoding='utf-8') as fh:
        json.dump(out, fh, ensure_ascii=False, indent=2)

    total = sum(len(b['processes']) for b in out)
    print(f"OK  {len(out)} fichiers, {total} process extraits -> {OUT_FILE}")
    for b in out:
        print(f"  - {b['file']:50s} {b['key_user']:35s} {len(b['processes'])} process")


if __name__ == '__main__':
    main()
