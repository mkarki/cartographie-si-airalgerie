"""
Genere REFERENTIEL_GLOBAL_A_VALIDER.xlsx :
fichier consolide de validation pour la secretaire / direction.

Sources :
  - cartographie_si/db.sqlite3 (cartography_structure, cartography_system,
    cartography_questionnaire)
  - LISTE_KEY_USERS (1).xlsx (referentiel emails valide secretaire)

Contenu :
  Feuille 1 — Lisez-moi
  Feuille 2 — Hierarchie (DG > Divisions > Directions + Responsables)
  Feuille 3 — Systemes (38 systemes : rattachement direction/division/criticite)
  Feuille 4 — Key Users (62 personnes : nom, prenom, email, role)
  Feuille 5 — Systeme <-> Key Users (matrice de rattachement)
  Feuille 6 — Diagnostic & ecarts a arbitrer
"""
import os
import re
import sqlite3
from collections import OrderedDict

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

BASE = os.path.dirname(os.path.abspath(__file__))
DB = os.path.join(BASE, 'db.sqlite3')
XLSX_KU = '/Users/mohamedamine/Air Algérie/LISTE_KEY_USERS (1).xlsx'
OUT = os.path.join(BASE, 'REFERENTIEL_GLOBAL_A_VALIDER.xlsx')

# Styles
H_FILL = PatternFill('solid', fgColor='1F4E78')
H_FONT = Font(bold=True, color='FFFFFF', size=11)
SUB_FILL = PatternFill('solid', fgColor='B4C7E7')
SUB_FONT = Font(bold=True, size=10)
TODO_FILL = PatternFill('solid', fgColor='FFF2CC')
WARN_FILL = PatternFill('solid', fgColor='F8CBAD')
OK_FILL = PatternFill('solid', fgColor='C6EFCE')
THIN = Side(border_style='thin', color='999999')
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
LEFT = Alignment(horizontal='left', vertical='center', wrap_text=True)
CENTER = Alignment(horizontal='center', vertical='center', wrap_text=True)


def style_header(ws, row, ncols):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = H_FILL
        cell.font = H_FONT
        cell.alignment = CENTER
        cell.border = BORDER
    ws.row_dimensions[row].height = 28


def set_widths(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


# -----------------------------------------------------------------------------
# Lecture des sources
# -----------------------------------------------------------------------------

def load_structures():
    conn = sqlite3.connect(DB)
    cur = conn.cursor()
    cur.execute('''
        SELECT s1.code, s1.name, s1.level, s1.responsable, s2.code as parent
        FROM cartography_structure s1
        LEFT JOIN cartography_structure s2 ON s1.parent_id = s2.id
    ''')
    rows = cur.fetchall()
    conn.close()
    return rows


def load_systems():
    conn = sqlite3.connect(DB)
    cur = conn.cursor()
    cur.execute('''
        SELECT s.code, s.name, s.vendor, s.criticality, st.code, st.name
        FROM cartography_system s
        LEFT JOIN cartography_structure st ON s.structure_id = st.id
        ORDER BY st.code, s.name
    ''')
    rows = cur.fetchall()
    conn.close()
    return rows


# Corrections secretaire (xlsx) appliquees UNIQUEMENT en lecture pour le
# referentiel. La BDD cartographie n'est PAS modifiee.
# Cle = forme exacte trouvee en BDD ; valeur = forme corrigee secretaire.
NAME_FIXES_FROM_XLSX = {
    'AIT MEZIANE Amar':         'AIT MEZIANE Ahcene',
    'BELKACEMI Mohand':         'BELKACEMI Mohand Said',
    'BENMADI Soumeya':          'BENMADI Soumia',
    'BOUTOUT Allaeddine':       'BOUTOUT Allaa Eddine',
    'BOUKAIOU Ahlem':           'BOUTOUT Allaa Eddine',
    'KARI Fateh':               'KARRI Fateh',
    'SAFAR ZITOUN Naim':        'SAFARZITOUN Naim',
    'BESLAMA Lamya':            'BENSALEM Lamya',
    'SMALAH Yasmine':           'SMAALLAH Yasmine',
    'AZEROUZL Salima':          'AZEROUAL Salima',
    'CHENOUFI Sid Ali':         'CHENNOUFI Sid Ali',
    'ESSEMINIA Adnan':          'ESSEMIANI Adnan',
    'SALAH ROUANA Med':         'SALAH ROUANA Mohamed',
    'BOUABDALLAH Akram':        'BOUABDALLAH Mohamed Akram',
    'BENNOUAR A.':              'BENNOUAR Adnane',
    'BENMOUFFOK E.':            'BENMOUFFOK El Hadi',
    'CHABANE A.':               'CHABANE Amel',
    'MAZARI A.':                'MAZARI Assia',
    'LADJICI A.':               'LADJICI Fatima',
    'BOUGUEZIZ':                'BOUGUEZIZ Samir',
    'FERRADJI':                 'FERRADJI Farid',
    'SAMEUR':                   'SAMEUR Yacine',
    'KARI':                     'KARRI Fateh',
    'BOUCHIK':                  'BOUCHIK Mounir',
    'BACHA':                    'BACHA Amine',
    'FODILI':                   'FODILI Mohamed Yacine',
    'BELDJERDI':                'BELDJERDI Zakaria',
    'AKKACHA':                  'AKKACHA Mohamed Amine',
    'SAAD':                     'SAAD Nassima',
    'SAAD Nasima':              'SAAD Nassima',
    'HALISSE':                  'HALISSE Abderrahim',
    'LAIDANI':                  'LAIDANI Zakaria',
    'FAIDI Foued':              'FAIDI Fouad',
    'HASSISSENE':               'HASSISSENE Chemseddine',
    'BENMADI':                  'BENMADI Soumia',
    'BENAOUICHA':               'BENAOUICHA Hanane',
    'GACIMI':                   'GACIMI Mohamed',
    'LAKAMA Abdallah':          'LAKAMA Abdellah',
}


def fix_names_in_text(s: str) -> str:
    """Applique les corrections du xlsx sur une chaine type
    'NOM1 (DIR), NOM2, NOM3 (DIR2)'. Tri par longueur decroissante pour
    eviter qu'une cle courte ne casse une forme deja longue."""
    if not s:
        return s
    out = s
    for old in sorted(NAME_FIXES_FROM_XLSX, key=len, reverse=True):
        new = NAME_FIXES_FROM_XLSX[old]
        if old == new:
            continue
        # On remplace uniquement si le nom apparait suivi d'un separateur
        # (virgule, parenthese, fin de chaine) — pas s'il est deja suivi
        # d'un autre mot (= deja la forme longue).
        pattern = re.compile(
            r'(?<![A-Za-zÀ-ÿ])' + re.escape(old) + r'(?=\s*(?:,|\(|/|$))'
        )
        out = pattern.sub(new, out)
    return out


def load_questionnaires():
    """Renvoie liste de (system_name, key_users, key_users_backup) sans les
    versions 'Questions Techniques —', avec corrections xlsx appliquees en
    memoire."""
    conn = sqlite3.connect(DB)
    cur = conn.cursor()
    cur.execute('''
        SELECT system_name, key_users, key_users_backup
        FROM cartography_questionnaire
        WHERE system_name NOT LIKE 'Questions Techniques%'
        ORDER BY system_name
    ''')
    rows = [(s, fix_names_in_text(k or ''), fix_names_in_text(b or ''))
            for s, k, b in cur.fetchall()]
    conn.close()
    return rows


def load_key_users_xlsx():
    """Renvoie liste : [(num, nom_prenom, email, direction, role, systemes)]"""
    wb = load_workbook(XLSX_KU)
    ws = wb.active
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] is None:
            continue
        rows.append(row)
    return rows


# -----------------------------------------------------------------------------
# Construction du classeur
# -----------------------------------------------------------------------------

def sheet_lisez_moi(wb):
    ws = wb.create_sheet('1. Lisez-moi')
    ws['A1'] = "Référentiel global SI Air Algérie — à valider"
    ws['A1'].font = Font(bold=True, size=16, color='1F4E78')

    ws['A3'] = "Objet"
    ws['A3'].font = Font(bold=True, size=12)
    ws['A4'] = (
        "Ce classeur consolide le référentiel utilisé par le projet de cartographie SI :\n"
        "  • Hiérarchie organisationnelle (Direction Générale → Divisions → Directions)\n"
        "  • Liste des systèmes recensés et leur direction de rattachement\n"
        "  • Liste des key users référents (nom, prénom, email)\n"
        "  • Matrice systèmes ↔ key users (qui est responsable de quel système)\n"
        "\n"
        "L'objectif est d'avoir UN SEUL référentiel partagé, validé par la "
        "direction, qui servira de source de vérité pour tous les livrables "
        "(mails, dashboards, fiches, ETL).")
    ws['A4'].alignment = Alignment(wrap_text=True, vertical='top')
    ws.row_dimensions[4].height = 130

    ws['A6'] = "Comment compléter ce fichier"
    ws['A6'].font = Font(bold=True, size=12)
    ws['A7'] = (
        "1. Feuille 2 (Hiérarchie) : valider/corriger le NOM des responsables et compléter leur EMAIL (cellules jaunes).\n"
        "2. Feuille 3 (Systèmes) : valider la direction de rattachement de chaque système.\n"
        "3. Feuille 4 (Key Users) : valider les noms et emails (déjà renseignés).\n"
        "4. Feuille 5 (Systèmes ↔ Key Users) : valider que les bons key users sont bien rattachés à chaque système.\n"
        "5. Feuille 6 (Diagnostic) : trancher les écarts détectés entre les sources.")
    ws['A7'].alignment = Alignment(wrap_text=True, vertical='top')
    ws.row_dimensions[7].height = 110

    ws['A9'] = "Légende des couleurs"
    ws['A9'].font = Font(bold=True, size=12)
    ws['A10'] = "Cellule jaune"
    ws['A10'].fill = TODO_FILL
    ws['B10'] = "À compléter / valider"
    ws['A11'] = "Cellule orange"
    ws['A11'].fill = WARN_FILL
    ws['B11'] = "Écart détecté entre sources — à arbitrer"
    ws['A12'] = "Cellule verte"
    ws['A12'].fill = OK_FILL
    ws['B12'] = "Validé / cohérent toutes sources"

    ws.column_dimensions['A'].width = 24
    ws.column_dimensions['B'].width = 90


def sheet_hierarchie(wb, structures):
    ws = wb.create_sheet('2. Hiérarchie')
    ws.append([
        '#', 'Code', 'Nom complet', 'Niveau',
        'Rattachement (parent)', 'Responsable', 'Email responsable (à compléter)',
        'Validé ?'
    ])
    style_header(ws, 1, 8)

    # Tri logique : DG, puis chaque Division, puis ses Directions
    parent_order = {'DG': 0, 'DAG': 1, 'DC': 2, 'DIVEX': 3}
    other_directions_under_dg = []

    by_parent = {}
    dg = None
    for code, name, level, resp, parent in structures:
        if level == 'DG':
            dg = (code, name, level, resp, parent)
        else:
            by_parent.setdefault(parent, []).append((code, name, level, resp, parent))

    rows_ordered = []
    if dg:
        rows_ordered.append(dg)

    # Divisions sous DG d'abord, puis directions sous DG
    for child in sorted(by_parent.get('DG', []), key=lambda x: (x[2] != 'Division', x[0])):
        rows_ordered.append(child)
        if child[2] == 'Division':
            for sub in sorted(by_parent.get(child[0], []), key=lambda x: x[0]):
                rows_ordered.append(sub)

    # Emails des responsables connus
    KNOWN_EMAILS = {
        'BOUTEMADJA SAMY':       'boutemadja.samy@airalgerie.dz',
        'KHELFI REDOUANE':       'khelfi.redouane@airalgerie.dz',
        'BENBOUAZIZ LYES':       'benbouaziz.lyes@airalgerie.dz',
        'HACHELAF MOURAD':       'hachelaf.mourad@airalgerie.dz',
        'KRAIMECHE ABDELKRIM':   'kraimeche.abdelkrim@airalgerie.dz',
        'MERDACI ADEL':          'merdaci.adel@airalgerie.dz',
        'FAIDI FOUAD':           'faidi.fouad@airalgerie.dz',
    }

    for i, (code, name, level, resp, parent) in enumerate(rows_ordered, 1):
        email = KNOWN_EMAILS.get((resp or '').strip().upper(), '')
        ws.append([
            i, code, name, level,
            parent or '—', resp or '(à désigner)',
            email,
            ''
        ])
        r = ws.max_row
        # Cellule email = jaune si vide
        if not email:
            ws.cell(row=r, column=7).fill = TODO_FILL
        else:
            ws.cell(row=r, column=7).fill = OK_FILL
        if not resp:
            ws.cell(row=r, column=6).fill = TODO_FILL
        ws.cell(row=r, column=8).fill = TODO_FILL
        # Bordures + alignement
        for c in range(1, 9):
            cell = ws.cell(row=r, column=c)
            cell.border = BORDER
            cell.alignment = LEFT if c in (3, 6, 7) else CENTER
        # Mettre en gras les Divisions
        if level in ('DG', 'Division'):
            for c in range(1, 9):
                ws.cell(row=r, column=c).font = Font(bold=True)
                if level == 'DG':
                    ws.cell(row=r, column=c).fill = PatternFill('solid', fgColor='FCE4D6')
                else:
                    ws.cell(row=r, column=c).fill = SUB_FILL

    set_widths(ws, [4, 8, 50, 12, 14, 28, 38, 10])
    ws.freeze_panes = 'A2'


def sheet_systemes(wb, systems):
    ws = wb.create_sheet('3. Systèmes')
    ws.append([
        '#', 'Code', 'Nom du système', 'Éditeur / Vendor',
        'Criticité', 'Direction de rattachement', 'Division parente',
        'Validé ?', 'Commentaire'
    ])
    style_header(ws, 1, 9)

    # Map direction -> division parente
    parent_map = {}
    for code, name, level, resp, parent in load_structures():
        if level == 'Direction' and parent:
            parent_map[code] = parent

    for i, (code, name, vendor, crit, struct_code, struct_name) in enumerate(systems, 1):
        division = parent_map.get(struct_code, struct_code) if struct_code else '—'
        ws.append([
            i, code or '', name, vendor,
            crit, struct_code or '—', division,
            '', ''
        ])
        r = ws.max_row
        # Couleur criticite
        crit_colors = {'CRITIQUE': 'F8CBAD', 'HAUTE': 'FFE699', 'MOYENNE': 'D9E1F2', 'BASSE': 'E2EFDA'}
        if crit in crit_colors:
            ws.cell(row=r, column=5).fill = PatternFill('solid', fgColor=crit_colors[crit])
        ws.cell(row=r, column=8).fill = TODO_FILL
        ws.cell(row=r, column=9).fill = TODO_FILL
        for c in range(1, 10):
            cell = ws.cell(row=r, column=c)
            cell.border = BORDER
            cell.alignment = LEFT if c in (3, 4, 9) else CENTER

    set_widths(ws, [4, 14, 42, 26, 12, 14, 12, 10, 35])
    ws.freeze_panes = 'A2'


def sheet_key_users(wb, ku_rows):
    ws = wb.create_sheet('4. Key Users')
    ws.append([
        '#', 'Nom & Prénom', 'Email', 'Direction(s)', 'Rôle',
        'Système(s) couvert(s)', 'Validé ?', 'Commentaire'
    ])
    style_header(ws, 1, 8)
    for row in ku_rows:
        num, name, email, direction, role, systems = row[0], row[1], row[2], row[3], row[4], row[5]
        ws.append([num, name, email, direction, role, systems, '', ''])
        r = ws.max_row
        ws.cell(row=r, column=3).fill = OK_FILL if email else TODO_FILL
        ws.cell(row=r, column=7).fill = TODO_FILL
        ws.cell(row=r, column=8).fill = TODO_FILL
        for c in range(1, 9):
            cell = ws.cell(row=r, column=c)
            cell.border = BORDER
            cell.alignment = LEFT if c in (2, 6, 8) else CENTER
    set_widths(ws, [4, 30, 38, 16, 18, 60, 10, 35])
    ws.freeze_panes = 'A2'


def sheet_systeme_keyusers(wb, questionnaires):
    ws = wb.create_sheet('5. Systèmes ↔ Key Users')
    ws.append([
        '#', 'Système (cartographie)', 'Key User Principal',
        'Key User Backup', 'Validé ?', 'Commentaire'
    ])
    style_header(ws, 1, 6)

    # Filtrage : on regroupe les sous-modules Altea/AIMS sous un seul item
    GROUP_MAP = {
        'Altéa Réservation': 'Suite Altéa Amadeus',
        'Altéa Ticketing':   'Suite Altéa Amadeus',
        'Altéa Inventory':   'Suite Altéa Amadeus',
        'Altéa DCS':         'Suite Altéa Amadeus',
        'Altéa RMS':         'Suite Altéa Amadeus',
        'Amadeus Data Feeds':'Suite Altéa Amadeus',
    }
    seen = set()
    rows = []
    for sys_name, kup, kub in questionnaires:
        norm = GROUP_MAP.get(sys_name, sys_name)
        if norm in seen:
            continue
        seen.add(norm)
        rows.append((norm, kup, kub))

    rows.sort(key=lambda x: x[0].lower())

    for i, (sys_name, kup, kub) in enumerate(rows, 1):
        ws.append([i, sys_name, kup or '—', kub or '—', '', ''])
        r = ws.max_row
        ws.cell(row=r, column=5).fill = TODO_FILL
        ws.cell(row=r, column=6).fill = TODO_FILL
        # Marquer en orange si "À désigner"
        if 'désigner' in (kup or '').lower():
            ws.cell(row=r, column=3).fill = WARN_FILL
        for c in range(1, 7):
            cell = ws.cell(row=r, column=c)
            cell.border = BORDER
            cell.alignment = LEFT if c in (2, 3, 4, 6) else CENTER
    set_widths(ws, [4, 38, 50, 50, 10, 35])
    ws.freeze_panes = 'A2'


def sheet_diagnostic(wb, ku_rows, questionnaires):
    """Detecte les ecarts entre la base cartographique et le xlsx KU."""
    ws = wb.create_sheet('6. Diagnostic')
    ws.append(['#', 'Catégorie', 'Élément', 'Source 1 (cartographie)', 'Source 2 (xlsx secrétaire)', 'À arbitrer'])
    style_header(ws, 1, 6)

    rows = []

    # 1. Ecarts noms entre DB cartography_questionnaire et xlsx KU
    # On extrait nom famille du xlsx
    xlsx_names = {row[1].split()[0].upper(): row[1] for row in ku_rows}
    db_names = set()
    for sys_name, kup, kub in questionnaires:
        for chunk in (kup or '').split(',') + (kub or '').split(','):
            n = chunk.strip().split('(')[0].strip()
            if n and n != '—' and 'désigner' not in n.lower():
                db_names.add(n)

    diffs_name = []
    for db_name in db_names:
        last = db_name.split()[0].upper()
        if last in xlsx_names:
            xlsx_n = xlsx_names[last]
            # Normaliser : compare en ignorant casse + espaces redondants
            norm_db = ' '.join(db_name.upper().split())
            norm_xl = ' '.join(xlsx_n.upper().split())
            if norm_db != norm_xl:
                diffs_name.append((db_name, xlsx_n))

    for db_name, xlsx_name in sorted(diffs_name):
        rows.append(('Nom', f'{db_name.split()[0]}', db_name, xlsx_name, ''))

    # 2. Personnes presentes dans le xlsx mais pas dans la DB (jamais vues)
    db_lastnames = {n.split()[0].upper() for n in db_names}
    xlsx_lastnames = {row[1].split()[0].upper(): row[1] for row in ku_rows}
    only_xlsx = sorted(set(xlsx_lastnames) - db_lastnames)
    for last in only_xlsx:
        rows.append(('Présence', last, '(absent)', xlsx_lastnames[last], ''))

    # 3. Personnes presentes dans la DB mais pas dans le xlsx
    only_db = sorted(db_lastnames - set(xlsx_lastnames))
    for last in only_db:
        # Recuperer le nom complet
        full = next((n for n in db_names if n.split()[0].upper() == last), last)
        rows.append(('Présence', last, full, '(absent)', ''))

    # 4. Systemes avec key user "À désigner"
    no_ku = sorted({sys for sys, kup, _ in questionnaires
                    if 'désigner' in (kup or '').lower()})
    for s in no_ku:
        rows.append(('Système sans KU', s, '(à désigner)', '', 'À désigner par la direction'))

    # 5. Directions sans email responsable (depuis DB)
    structs = load_structures()
    KNOWN = {'BOUTEMADJA SAMY', 'KHELFI REDOUANE', 'BENBOUAZIZ LYES',
             'HACHELAF MOURAD', 'KRAIMECHE ABDELKRIM', 'MERDACI ADEL',
             'FAIDI FOUAD'}
    for code, name, level, resp, parent in structs:
        if level in ('Direction',) and resp and resp.strip().upper() not in KNOWN:
            rows.append(('Email manquant', f'{code} ({name})', resp, '', 'Email responsable à fournir'))

    for i, r in enumerate(rows, 1):
        ws.append([i, r[0], r[1], r[2], r[3], r[4]])
        rr = ws.max_row
        ws.cell(row=rr, column=6).fill = TODO_FILL
        # Marquer en orange tous les ecarts "Nom"
        if r[0] == 'Nom':
            for c in range(2, 6):
                ws.cell(row=rr, column=c).fill = WARN_FILL
        for c in range(1, 7):
            cell = ws.cell(row=rr, column=c)
            cell.border = BORDER
            cell.alignment = LEFT if c in (3, 4, 5, 6) else CENTER

    set_widths(ws, [4, 18, 28, 34, 34, 38])
    ws.freeze_panes = 'A2'


# -----------------------------------------------------------------------------
# Main
# -----------------------------------------------------------------------------

def main():
    wb = Workbook()
    # Supprimer feuille par defaut
    wb.remove(wb.active)

    structures = load_structures()
    systems = load_systems()
    questionnaires = load_questionnaires()
    ku_rows = load_key_users_xlsx()

    sheet_lisez_moi(wb)
    sheet_hierarchie(wb, structures)
    sheet_systemes(wb, systems)
    sheet_key_users(wb, ku_rows)
    sheet_systeme_keyusers(wb, questionnaires)
    sheet_diagnostic(wb, ku_rows, questionnaires)

    wb.save(OUT)
    print(f'OK  {OUT}')
    print(f'    Feuilles : Lisez-moi, Hierarchie ({len(structures)} entites), '
          f'Systemes ({len(systems)}), Key Users ({len(ku_rows)}), '
          f'Systemes-KU, Diagnostic')


if __name__ == '__main__':
    main()
